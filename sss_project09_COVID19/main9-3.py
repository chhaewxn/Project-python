from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import requests
import re
import datetime

import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))


url = "https://apis.data.go.kr/1352000/ODMS_COVID_04/callCovid04Api?serviceKey=xdF4c2%2FejrJEYFVOOsmUXo4ueNyAunC75B52bGYD%2FO4%2F7GigXR8FTq5hVkarOLnLoNnFZKMS4e8ndavGalT1wA%3D%3D&pageNo=1&numOfRows=500&apiType=xml"
api_key = "xdF4c2%2FejrJEYFVOOsmUXo4ueNyAunC75B52bGYD%2FO4%2F7GigXR8FTq5hVkarOLnLoNnFZKMS4e8ndavGalT1wA%3D%3D"


def get_today_covid19(url, api_key):
    now = datetime.datetime.now()
    yyyymmdd = now.strftime('%Y%m%d')

    page_no = "&pageNo=1&numOfRows=30&"
    today = "startCreateDt=" + yyyymmdd + "&endCreateDt=" + yyyymmdd

    response = requests.get(url + api_key + page_no + today)

    gubun = re.findall(r'<gubun>(.+?)</gubun>', response.text)
    incDec = re.findall(r'<incDec>(.+?)</incDec>', response.text)

    return yyyymmdd, gubun, incDec


날짜, 지역, 확진자수 = get_today_covid19(url, api_key)

try:
    wb = load_workbook("코로나확진자보고서.xlsx", data_only=True)
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active


ws.cell(row=1, column=1).value = 날짜 + "일 코로나 현황"
ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(확진자수))
ws.cell(row=1, column=1).alignment = Alignment(
    horizontal="center", vertical="center")

for i in range(len(확진자수)):
    ws.cell(row=2, column=i+1).value = 지역[i]
    ws.cell(row=3, column=i+1).value = int(확진자수[i])


wb.save("코로나확진자보고서.xlsx")
